# Gather and combine .xlsx and .csv files and combine them into a single worksheet/workbook.
# Michael Woolcott

import openpyxl
import csv
import os


def remove_illegal_characters(value):
    illegal_characters = ['\x00', '\x01', '\x02', '\x03', '\x04','\x05', '\x06', '\x07', '\x08', '\x0B', 
                          '\x0C', '\x0E', '\x0F', '\x10', '\x11', '\x12', '\x13', '\x14', '\x15', '\x16', 
                          '\x17', '\x18', '\x19', '\x1A','\x1B', '\x1C', '\x1D', '\x1E', '\x1F', ':', '*', 
                          '?', '[', ']', '/' '\\']

    for char in illegal_characters:
        value = value.replace(char, '')
    return value

def combine_files():
    current_directory = os.getcwd()
    combined_workbook = openpyxl.Workbook()
    combined_sheet = combined_workbook.active

    for filename in os.listdir(current_directory):
        if filename.endswith(".csv") or filename.endswith(".xlsx"):
            try:
                with open(filename, 'rb') as file:
                    byte_content = file.read()
                    decoded_content = byte_content.decode('latin-1')

                cleaned_content = decoded_content.replace('\x00', '')

                if filename.endswith(".csv"):
                    reader = csv.reader(cleaned_content.splitlines())
                    data = list(reader)
                elif filename.endswith(".xlsx"):
                    workbook = openpyxl.load_workbook(filename)
                    sheet = workbook.active
                    data = [list(row) for row in sheet.iter_rows(values_only=True)]

                for row in data:
                    cleaned_row = [remove_illegal_characters(str(value)) for value in row]
                    combined_sheet.append(cleaned_row)
            except KeyError as e:
                print(f"Skipping file '{filename}' due to error: {e}")

    combined_workbook.save("combined_data.xlsx")


combine_files()         


